"""
Brivia Pricing PRO · v4.0
Multi-page · SQLite · Analytics · Excel Import
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import sqlite3
import json
import random
import io
import os
from datetime import datetime, timedelta
from decimal import Decimal, getcontext
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

getcontext().prec = 20

# ==================== PAGE CONFIG ====================

st.set_page_config(
    page_title="Brivia Pricing PRO",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== CONSTANTS ====================

DB_PATH = Path(__file__).parent / "brivia_pricing.db"
LOGO_URL = "https://cdn.prod.website-files.com/65c2dcb4330facd527e06bdd/6619a7597575e945de440959_brivia_group.svg"

TIPOS_CONTRATO = [
    "Fee Mensal (Recorrente)",
    "Projeto (Escopo Fechado)",
    "Sustentação",
    "Consultoria Estratégica"
]
TIPOS_SERVICO = [
    "Tecnologia / Dev",
    "Design / UX",
    "Dados / Analytics",
    "Marketing / Growth",
    "Estratégia Digital"
]
STATUS_LIST = ["Rascunho", "Em Análise", "Aprovada", "Rejeitada", "Em Execução", "Concluída", "Cancelada"]
STATUS_COLORS = {
    "Rascunho":     "#6b7280",
    "Em Análise":   "#3b82f6",
    "Aprovada":     "#10b981",
    "Rejeitada":    "#ef4444",
    "Em Execução":  "#f59e0b",
    "Concluída":    "#c58f3d",
    "Cancelada":    "#9ca3af",
}
MAPEAMENTO_OFERTAS = {
    "Tecnologia / Dev":   "Tech Squad",
    "Design / UX":        "Design Studio",
    "Dados / Analytics":  "Data Intelligence",
    "Marketing / Growth": "Growth Lab",
    "Estratégia Digital": "Strategy Hub",
}

COMISSAO_NB       = [0, 5, 7, 10, 12, 15]
COMISSAO_PARCEIROS = [0, 5, 10, 15, 20, 25]
GROSS_MARGIN_ALVO  = 45.0
IMPOSTO_PADRAO_PCT = 14.25
BENEFICIOS_MENSAIS = Decimal("1190")
HORAS_MES          = Decimal("170")

_P13 = Decimal("100") / Decimal("12")
_PFE = Decimal("11.1110833333333")
ENCARGOS_FPA = {
    "fgts_base":      Decimal("8.00"),
    "fgts_s_13":      _P13 * Decimal("0.08"),
    "fgts_s_ferias":  _PFE * Decimal("0.08"),
    "inss_base":      Decimal("26.30"),
    "inss_s_13":      _P13 * Decimal("0.263"),
    "inss_s_ferias":  _PFE * Decimal("0.263"),
    "prov_13":        _P13,
    "prov_ferias":    _PFE,
    "aviso_previo":   Decimal("1.32"),
    "aux_doenca":     Decimal("0.55"),
    "rescisao":       Decimal("2.57"),
}
FATOR_ENCARGOS = Decimal("1") + (sum(ENCARGOS_FPA.values()) / Decimal("100"))

BASE_SALARIAL = {
    "Desenvolvedor Fullstack": {
        "1. Júnior": {"mercado": 4500,  "minima": 3800,  "brivia": 5000},
        "2. Pleno":  {"mercado": 7500,  "minima": 6500,  "brivia": 8500},
        "3. Sênior": {"mercado": 11000, "minima": 9500,  "brivia": 13000},
        "4. Líder":  {"mercado": 14000, "minima": 12000, "brivia": 16000},
        "5. Head":   {"mercado": 18000, "minima": 15000, "brivia": 22000},
    },
    "Desenvolvedor Frontend": {
        "1. Júnior": {"mercado": 4000,  "minima": 3500,  "brivia": 4500},
        "2. Pleno":  {"mercado": 6500,  "minima": 5500,  "brivia": 7500},
        "3. Sênior": {"mercado": 10000, "minima": 8500,  "brivia": 12000},
        "4. Líder":  {"mercado": 13000, "minima": 11000, "brivia": 15000},
        "5. Head":   {"mercado": 17000, "minima": 14000, "brivia": 20000},
    },
    "Desenvolvedor Backend": {
        "1. Júnior": {"mercado": 4500,  "minima": 3800,  "brivia": 5000},
        "2. Pleno":  {"mercado": 7500,  "minima": 6500,  "brivia": 8500},
        "3. Sênior": {"mercado": 11500, "minima": 10000, "brivia": 13500},
        "4. Líder":  {"mercado": 14500, "minima": 12500, "brivia": 17000},
        "5. Head":   {"mercado": 19000, "minima": 16000, "brivia": 23000},
    },
    "Cientista de Dados": {
        "1. Júnior": {"mercado": 5500,  "minima": 4500,  "brivia": 6500},
        "2. Pleno":  {"mercado": 9000,  "minima": 7500,  "brivia": 10500},
        "3. Sênior": {"mercado": 13500, "minima": 11500, "brivia": 16000},
        "4. Líder":  {"mercado": 17000, "minima": 14500, "brivia": 20000},
        "5. Head":   {"mercado": 22000, "minima": 18000, "brivia": 26000},
    },
    "Engenheiro de Dados": {
        "1. Júnior": {"mercado": 5000,  "minima": 4200,  "brivia": 5800},
        "2. Pleno":  {"mercado": 8500,  "minima": 7200,  "brivia": 10000},
        "3. Sênior": {"mercado": 13000, "minima": 11000, "brivia": 15500},
        "4. Líder":  {"mercado": 16500, "minima": 14000, "brivia": 19500},
        "5. Head":   {"mercado": 21000, "minima": 17500, "brivia": 25000},
    },
    "UX Designer": {
        "1. Júnior": {"mercado": 4000,  "minima": 3200,  "brivia": 4500},
        "2. Pleno":  {"mercado": 6500,  "minima": 5500,  "brivia": 7500},
        "3. Sênior": {"mercado": 9500,  "minima": 8000,  "brivia": 11000},
        "4. Líder":  {"mercado": 12500, "minima": 10500, "brivia": 14500},
        "5. Head":   {"mercado": 16000, "minima": 13500, "brivia": 19000},
    },
    "UI Designer": {
        "1. Júnior": {"mercado": 3800,  "minima": 3000,  "brivia": 4200},
        "2. Pleno":  {"mercado": 6000,  "minima": 5000,  "brivia": 7000},
        "3. Sênior": {"mercado": 9000,  "minima": 7500,  "brivia": 10500},
        "4. Líder":  {"mercado": 12000, "minima": 10000, "brivia": 14000},
        "5. Head":   {"mercado": 15500, "minima": 13000, "brivia": 18000},
    },
    "Product Manager": {
        "1. Júnior": {"mercado": 5000,  "minima": 4200,  "brivia": 5800},
        "2. Pleno":  {"mercado": 8500,  "minima": 7200,  "brivia": 10000},
        "3. Sênior": {"mercado": 13000, "minima": 11000, "brivia": 15000},
        "4. Líder":  {"mercado": 17000, "minima": 14500, "brivia": 20000},
        "5. Head":   {"mercado": 22000, "minima": 18500, "brivia": 26000},
    },
    "Tech Lead": {
        "3. Sênior": {"mercado": 14000, "minima": 12000, "brivia": 16500},
        "4. Líder":  {"mercado": 18000, "minima": 15500, "brivia": 21000},
        "5. Head":   {"mercado": 24000, "minima": 20000, "brivia": 28000},
    },
    "DevOps/SRE": {
        "1. Júnior": {"mercado": 5000,  "minima": 4200,  "brivia": 5800},
        "2. Pleno":  {"mercado": 8500,  "minima": 7200,  "brivia": 10000},
        "3. Sênior": {"mercado": 13000, "minima": 11000, "brivia": 15500},
        "4. Líder":  {"mercado": 17000, "minima": 14500, "brivia": 20000},
        "5. Head":   {"mercado": 22000, "minima": 18500, "brivia": 26000},
    },
    "Analista de Marketing": {
        "1. Júnior": {"mercado": 3500,  "minima": 2800,  "brivia": 4000},
        "2. Pleno":  {"mercado": 5500,  "minima": 4500,  "brivia": 6500},
        "3. Sênior": {"mercado": 8500,  "minima": 7000,  "brivia": 10000},
        "4. Líder":  {"mercado": 12000, "minima": 10000, "brivia": 14000},
        "5. Head":   {"mercado": 16000, "minima": 13500, "brivia": 19000},
    },
    "Scrum Master": {
        "2. Pleno":  {"mercado": 7000,  "minima": 5800,  "brivia": 8200},
        "3. Sênior": {"mercado": 10500, "minima": 9000,  "brivia": 12500},
        "4. Líder":  {"mercado": 14000, "minima": 12000, "brivia": 16500},
        "5. Head":   {"mercado": 18000, "minima": 15000, "brivia": 21000},
    },
}

COR_PRIMARIA  = "#c58f3d"
COR_ACENTO    = "#3b82f6"
COR_SUCESSO   = "#10b981"
COR_ALERTA    = "#f59e0b"
COR_ERRO      = "#ef4444"

CHART_COLORS = [COR_PRIMARIA, COR_ACENTO, COR_SUCESSO, COR_ALERTA, "#8b5cf6", "#ec4899", "#06b6d4"]

# ==================== DATABASE LAYER ====================

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS propostas (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at      TEXT    DEFAULT (datetime('now','localtime')),
                updated_at      TEXT    DEFAULT (datetime('now','localtime')),
                status          TEXT    DEFAULT 'Rascunho',
                cliente         TEXT    NOT NULL,
                segmento        TEXT,
                projeto         TEXT    NOT NULL,
                descricao       TEXT,
                tipo_contrato   TEXT,
                tipo_servico    TEXT,
                meses           INTEGER DEFAULT 12,
                imposto_pct     REAL    DEFAULT 14.25,
                comissao_nb     REAL    DEFAULT 0,
                comissao_parc   REAL    DEFAULT 0,
                gm_alvo         REAL    DEFAULT 45,
                regua_salarial  TEXT    DEFAULT 'mercado',
                custo_equipe    REAL    DEFAULT 0,
                custo_terceiros REAL    DEFAULT 0,
                custo_extras    REAL    DEFAULT 0,
                custo_total     REAL    DEFAULT 0,
                preco_venda     REAL    DEFAULT 0,
                fee_mensal      REAL    DEFAULT 0,
                v_impostos      REAL    DEFAULT 0,
                v_comissoes     REAL    DEFAULT 0,
                margem_bruta    REAL    DEFAULT 0,
                markup_pct      REAL    DEFAULT 0,
                headcount       INTEGER DEFAULT 0,
                responsavel     TEXT,
                obs             TEXT
            );

            CREATE TABLE IF NOT EXISTS proposta_equipe (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                proposta_id INTEGER NOT NULL,
                perfil      TEXT,
                nivel       TEXT,
                qtd         INTEGER DEFAULT 1,
                dedicacao   REAL    DEFAULT 100,
                salario_base REAL   DEFAULT 0,
                regua       TEXT,
                FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS proposta_terceiros (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                proposta_id INTEGER NOT NULL,
                descricao   TEXT,
                valor       REAL DEFAULT 0,
                FOREIGN KEY (proposta_id) REFERENCES propostas(id) ON DELETE CASCADE
            );
        """)


def db_count():
    with get_db() as conn:
        return conn.execute("SELECT COUNT(*) FROM propostas").fetchone()[0]


def db_get_all(filtros=None):
    query = "SELECT * FROM propostas"
    params = []
    conds  = []
    if filtros:
        if filtros.get("status"):
            conds.append("status = ?"); params.append(filtros["status"])
        if filtros.get("tipo_servico"):
            conds.append("tipo_servico = ?"); params.append(filtros["tipo_servico"])
        if filtros.get("tipo_contrato"):
            conds.append("tipo_contrato = ?"); params.append(filtros["tipo_contrato"])
        if filtros.get("cliente"):
            conds.append("cliente LIKE ?"); params.append(f"%{filtros['cliente']}%")
    if conds:
        query += " WHERE " + " AND ".join(conds)
    query += " ORDER BY created_at DESC"
    with get_db() as conn:
        return [dict(r) for r in conn.execute(query, params).fetchall()]


def db_get_one(pid):
    with get_db() as conn:
        p = conn.execute("SELECT * FROM propostas WHERE id=?", (pid,)).fetchone()
        eq = conn.execute("SELECT * FROM proposta_equipe WHERE proposta_id=?", (pid,)).fetchall()
        te = conn.execute("SELECT * FROM proposta_terceiros WHERE proposta_id=?", (pid,)).fetchall()
        if not p:
            return None, [], []
        return dict(p), [dict(e) for e in eq], [dict(t) for t in te]


def db_save(data, equipe, terceiros):
    with get_db() as conn:
        pid = data.get("id")
        fields = (
            data["status"], data["cliente"], data.get("segmento", ""),
            data["projeto"], data.get("descricao", ""),
            data["tipo_contrato"], data["tipo_servico"], data["meses"],
            data["imposto_pct"], data["comissao_nb"], data["comissao_parc"],
            data["gm_alvo"], data["regua_salarial"],
            data["custo_equipe"], data["custo_terceiros"], data["custo_extras"],
            data["custo_total"], data["preco_venda"], data["fee_mensal"],
            data["v_impostos"], data["v_comissoes"], data["margem_bruta"],
            data["markup_pct"], data["headcount"],
            data.get("responsavel", ""), data.get("obs", ""),
        )
        if pid:
            conn.execute("""
                UPDATE propostas SET
                    updated_at=datetime('now','localtime'), status=?, cliente=?, segmento=?,
                    projeto=?, descricao=?, tipo_contrato=?, tipo_servico=?, meses=?,
                    imposto_pct=?, comissao_nb=?, comissao_parc=?, gm_alvo=?, regua_salarial=?,
                    custo_equipe=?, custo_terceiros=?, custo_extras=?, custo_total=?,
                    preco_venda=?, fee_mensal=?, v_impostos=?, v_comissoes=?,
                    margem_bruta=?, markup_pct=?, headcount=?, responsavel=?, obs=?
                WHERE id=?
            """, (*fields, pid))
            conn.execute("DELETE FROM proposta_equipe WHERE proposta_id=?", (pid,))
            conn.execute("DELETE FROM proposta_terceiros WHERE proposta_id=?", (pid,))
        else:
            cur = conn.execute("""
                INSERT INTO propostas (
                    status, cliente, segmento, projeto, descricao,
                    tipo_contrato, tipo_servico, meses,
                    imposto_pct, comissao_nb, comissao_parc, gm_alvo, regua_salarial,
                    custo_equipe, custo_terceiros, custo_extras, custo_total,
                    preco_venda, fee_mensal, v_impostos, v_comissoes,
                    margem_bruta, markup_pct, headcount, responsavel, obs
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, fields)
            pid = cur.lastrowid

        for e in equipe:
            conn.execute(
                "INSERT INTO proposta_equipe (proposta_id,perfil,nivel,qtd,dedicacao,salario_base,regua) VALUES (?,?,?,?,?,?,?)",
                (pid, e["perfil"], e["nivel"], e["qtd"], e["dedicacao"], e["salario_base"], e["regua"]),
            )
        for t in terceiros:
            conn.execute(
                "INSERT INTO proposta_terceiros (proposta_id,descricao,valor) VALUES (?,?,?)",
                (pid, t["desc"], t["valor"]),
            )
        return pid


def db_delete(pid):
    with get_db() as conn:
        conn.execute("DELETE FROM propostas WHERE id=?", (pid,))


def db_update_status(pid, status):
    with get_db() as conn:
        conn.execute(
            "UPDATE propostas SET status=?, updated_at=datetime('now','localtime') WHERE id=?",
            (status, pid),
        )

# ==================== ARTIFICIAL DATA SEED ====================

_CLIENTES = [
    ("Itaú Unibanco", "Banco"), ("Nubank", "Fintech"),
    ("Magazine Luiza", "Varejo"), ("B3 - Bolsa de Valores", "Mercado de Capitais"),
    ("Ambev", "Bebidas/FMCG"), ("Petrobras", "Energia"),
    ("Grupo Globo", "Mídia"), ("Embraer", "Aeroespacial"),
    ("Santander Brasil", "Banco"), ("Vivo Telefônica", "Telecom"),
    ("Hapvida Intermédica", "Saúde"), ("Vale", "Mineração"),
    ("Cielo", "Meios de Pagamento"), ("XP Inc.", "Investimentos"),
    ("Mercado Livre", "E-commerce"), ("iFood", "Foodtech"),
    ("TOTVS", "ERP/Software"), ("Localiza Hertz", "Mobilidade"),
    ("Rede D'Or São Luiz", "Saúde"), ("Porto Seguro", "Seguros"),
]
_PROJETOS = {
    "Tecnologia / Dev": [
        "Plataforma Digital de Serviços", "Modernização de Core Bancário",
        "Squad Dedicado Full-Stack", "API Gateway e Microsserviços",
        "Aplicativo Mobile B2C", "Portal do Cliente 2.0",
        "Automação de Processos Críticos", "Migração para Cloud AWS",
        "Sistema de Onboarding Digital", "Plataforma de Pagamentos PIX",
    ],
    "Design / UX": [
        "Redesign de Experiência do Usuário", "Design System Corporativo",
        "Research & Discovery Sprint", "UX para App Mobile iOS/Android",
        "Identidade Visual Digital", "Auditoria de Acessibilidade UI",
    ],
    "Dados / Analytics": [
        "Data Lake Corporativo", "Dashboard Executivo BI",
        "Modelo Preditivo de Churn", "Data Governance Implementation",
        "Analytics em Tempo Real", "CDP - Customer Data Platform",
        "Engenharia de Features ML", "Data Mesh Architecture",
    ],
    "Marketing / Growth": [
        "Growth Hacking para Aquisição", "Performance Marketing Digital",
        "CRM e Automação de Marketing", "SEO e Estratégia de Conteúdo",
        "Retenção e Loyalty Program", "Social Commerce Strategy",
    ],
    "Estratégia Digital": [
        "Transformação Digital Corporativa", "Roadmap Tecnológico 2025-2027",
        "Assessment de Maturidade Digital", "Consultoria em Inovação Aberta",
        "Due Diligence Tecnológica", "OKRs e Gestão Ágil de Portfólio",
    ],
}
_RESPONSAVEIS = [
    "Ana Paula Ferreira", "Carlos Eduardo Lima", "Fernanda Souza",
    "Rafael Moreira", "Juliana Castro", "Bruno Alves",
    "Gabriela Nunes", "Thiago Ribeiro",
]
_PERFIS_POR_SERVICO = {
    "Tecnologia / Dev": [
        ("Desenvolvedor Fullstack", ["2. Pleno", "3. Sênior"]),
        ("Desenvolvedor Backend",   ["2. Pleno", "3. Sênior"]),
        ("Desenvolvedor Frontend",  ["2. Pleno", "3. Sênior"]),
        ("Tech Lead",               ["3. Sênior", "4. Líder"]),
        ("DevOps/SRE",              ["2. Pleno", "3. Sênior"]),
        ("Scrum Master",            ["2. Pleno", "3. Sênior"]),
    ],
    "Design / UX": [
        ("UX Designer",    ["2. Pleno", "3. Sênior"]),
        ("UI Designer",    ["2. Pleno", "3. Sênior"]),
        ("Product Manager",["2. Pleno", "3. Sênior"]),
    ],
    "Dados / Analytics": [
        ("Cientista de Dados", ["2. Pleno", "3. Sênior"]),
        ("Engenheiro de Dados",["2. Pleno", "3. Sênior"]),
        ("Desenvolvedor Backend",["2. Pleno", "3. Sênior"]),
    ],
    "Marketing / Growth": [
        ("Analista de Marketing", ["2. Pleno", "3. Sênior"]),
        ("UX Designer",           ["2. Pleno"]),
        ("Cientista de Dados",    ["2. Pleno"]),
    ],
    "Estratégia Digital": [
        ("Product Manager",    ["3. Sênior", "4. Líder"]),
        ("Cientista de Dados", ["3. Sênior"]),
        ("Tech Lead",          ["4. Líder"]),
    ],
}
_SERVICOS_TERC = [
    "Consultoria Cloud AWS", "Freelancer Design", "Parceiro QA",
    "Infra Azure", "Licença Figma Pro", "AWS Reserved Instances",
    "Consultoria Segurança", "Suporte Salesforce", "Auditoria Externa",
]


def _custo_item(salario, ded_pct, meses, qtd):
    s = Decimal(str(salario))
    d = Decimal(str(ded_pct)) / Decimal("100")
    return float((s * FATOR_ENCARGOS + BENEFICIOS_MENSAIS) * d * Decimal(str(meses)) * Decimal(str(qtd)))


def _preco_reverso(custo, gm, imp, com):
    div = 1 - (gm + imp + com) / 100
    if div <= 0:
        return custo * 2, 0, 0, 0
    p = custo / div
    return p, p * imp / 100, p * com / 100, p * gm / 100


def seed_data():
    if db_count() > 0:
        return
    random.seed(42)
    start = datetime(2024, 1, 15)
    end   = datetime(2026, 2, 28)
    span  = (end - start).days

    for _ in range(80):
        created = start + timedelta(days=random.randint(0, span))
        cliente, segmento  = random.choice(_CLIENTES)
        tipo_serv = random.choice(TIPOS_SERVICO)
        tipo_cont = random.choice(TIPOS_CONTRATO)
        projeto   = random.choice(_PROJETOS[tipo_serv])
        meses     = random.choice([3, 6, 9, 12, 12, 18, 24])
        imp       = random.choice([12.0, 13.5, 14.25, 15.0, 16.5])
        com_nb    = random.choice([0, 5, 7, 10])
        com_pa    = random.choice([0, 0, 5, 10])
        gm        = random.choice([35.0, 40.0, 45.0, 50.0])
        regua     = random.choice(["mercado", "mercado", "brivia", "minima"])
        resp      = random.choice(_RESPONSAVEIS)

        perfis_disp = _PERFIS_POR_SERVICO.get(tipo_serv, [])
        n = random.randint(2, min(5, len(perfis_disp)))
        selecionados = random.sample(perfis_disp, n)

        equipe = []
        custo_eq = 0
        headcount = 0
        for perfil, niveis in selecionados:
            nivel = random.choice(niveis)
            sal   = BASE_SALARIAL.get(perfil, {}).get(nivel, {}).get(regua, 8000)
            qtd   = random.choice([1, 1, 1, 2, 2, 3])
            ded   = random.choice([50, 75, 100, 100])
            custo_eq  += _custo_item(sal, ded, meses, qtd)
            headcount += qtd
            equipe.append({"perfil": perfil, "nivel": nivel, "qtd": qtd,
                           "dedicacao": ded, "salario_base": sal, "regua": regua})

        terc_list = []
        custo_te  = 0
        if random.random() > 0.65:
            for _ in range(random.randint(1, 2)):
                v = random.choice([5000, 8000, 12000, 15000, 20000, 30000])
                terc_list.append({"desc": random.choice(_SERVICOS_TERC), "valor": v})
                custo_te += v

        extras = {
            "viagens":      random.choice([0, 0, 500, 1000, 2000]),
            "software":     random.choice([0, 300, 500, 1000, 1500]),
            "infraestrutura": random.choice([0, 500, 1000, 2000, 3000]),
            "outros":       random.choice([0, 0, 300, 500]),
        }
        custo_ex    = sum(extras.values()) * meses
        custo_total = custo_eq + custo_te + custo_ex
        com_total   = com_nb + com_pa
        preco, v_imp, v_com, v_gm = _preco_reverso(custo_total, gm, imp, com_total)
        fee   = preco / meses
        mkp   = ((preco / custo_total) - 1) * 100 if custo_total else 0

        age_m = (datetime.now() - created).days / 30
        if age_m < 2:
            status = random.choices(["Rascunho", "Em Análise", "Aprovada"], weights=[20, 50, 30])[0]
        elif age_m < 6:
            status = random.choices(["Em Análise", "Aprovada", "Rejeitada", "Em Execução"], weights=[10, 30, 25, 35])[0]
        else:
            status = random.choices(["Aprovada", "Rejeitada", "Em Execução", "Concluída", "Cancelada"], weights=[10, 25, 20, 35, 10])[0]

        created_s = created.strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as conn:
            cur = conn.execute("""
                INSERT INTO propostas (
                    created_at, updated_at, status, cliente, segmento, projeto, descricao,
                    tipo_contrato, tipo_servico, meses, imposto_pct, comissao_nb, comissao_parc,
                    gm_alvo, regua_salarial, custo_equipe, custo_terceiros, custo_extras,
                    custo_total, preco_venda, fee_mensal, v_impostos, v_comissoes,
                    margem_bruta, markup_pct, headcount, responsavel, obs
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                created_s, created_s, status, cliente, segmento,
                projeto, f"Proposta gerada artificialmente — {projeto}",
                tipo_cont, tipo_serv, meses, imp, com_nb, com_pa,
                gm, regua, custo_eq, custo_te, custo_ex,
                custo_total, preco, fee, v_imp, v_com, v_gm,
                mkp, headcount, resp, "",
            ))
            pid = cur.lastrowid
            for e in equipe:
                conn.execute(
                    "INSERT INTO proposta_equipe (proposta_id,perfil,nivel,qtd,dedicacao,salario_base,regua) VALUES (?,?,?,?,?,?,?)",
                    (pid, e["perfil"], e["nivel"], e["qtd"], e["dedicacao"], e["salario_base"], e["regua"]),
                )
            for t in terc_list:
                conn.execute(
                    "INSERT INTO proposta_terceiros (proposta_id,descricao,valor) VALUES (?,?,?)",
                    (pid, t["desc"], t["valor"]),
                )

# ==================== CSS ====================

def inject_css():
    st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600&display=swap');

/* ── RESET ── */
*, *::before, *::after {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    box-sizing: border-box;
}}

.stApp {{
    background: #0d0d0d !important;
}}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"] {{
    background: #080808 !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
    width: 220px !important;
}}
section[data-testid="stSidebar"] > div {{
    padding: 0 !important;
}}

/* Sidebar buttons — nav items */
section[data-testid="stSidebar"] div.stButton > button {{
    width: 100% !important;
    background: transparent !important;
    border: none !important;
    color: rgba(255,255,255,0.45) !important;
    text-align: left !important;
    padding: 10px 20px !important;
    border-radius: 0 !important;
    font-size: 0.85rem !important;
    font-weight: 400 !important;
    letter-spacing: 0 !important;
    text-transform: none !important;
    transition: all 0.2s !important;
    box-shadow: none !important;
}}
section[data-testid="stSidebar"] div.stButton > button:hover {{
    background: rgba(255,255,255,0.04) !important;
    color: rgba(255,255,255,0.85) !important;
    transform: none !important;
    box-shadow: none !important;
}}

/* ── MAIN CONTENT ── */
.block-container {{
    padding: 2rem 2.5rem !important;
    max-width: 100% !important;
}}

/* ── TYPOGRAPHY ── */
h1 {{ font-size: 1.75rem !important; font-weight: 700 !important; color: #f9fafb !important; letter-spacing: -0.5px !important; margin-bottom: 4px !important; }}
h2 {{ font-size: 1.25rem !important; font-weight: 600 !important; color: #f9fafb !important; }}
h3 {{ font-size: 1rem !important;   font-weight: 600 !important; color: #f9fafb !important; }}
p, span, label, div {{ color: #9ca3af; }}

/* ── SCROLLBAR ── */
* {{ scrollbar-width: thin; scrollbar-color: #2d2d2d transparent; }}
*::-webkit-scrollbar {{ width: 5px; height: 5px; }}
*::-webkit-scrollbar-track {{ background: transparent; }}
*::-webkit-scrollbar-thumb {{ background: #2d2d2d; border-radius: 3px; }}

/* ── CARDS ── */
.pro-card {{
    background: #111111;
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 20px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s;
}}
.pro-card:hover {{ border-color: rgba(255,255,255,0.1); }}
.pro-card-header {{
    display: flex; align-items: center; gap: 10px;
    margin-bottom: 20px; padding-bottom: 16px;
    border-bottom: 1px solid rgba(255,255,255,0.05);
}}
.pro-card-title {{
    font-size: 0.9rem; font-weight: 600;
    color: #e5e7eb; letter-spacing: -0.2px;
}}
.card-accent-left {{
    border-left: 3px solid {COR_PRIMARIA};
}}

/* ── KPI CARDS ── */
.kpi-wrap {{
    background: #111111;
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 12px;
    padding: 20px 22px;
    position: relative;
    overflow: hidden;
}}
.kpi-wrap::after {{
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, var(--kpi-color, {COR_PRIMARIA}), transparent);
    opacity: 0.6;
}}
.kpi-label {{
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 1.2px;
    color: rgba(255,255,255,0.35);
    margin-bottom: 10px;
}}
.kpi-value {{
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 1.7rem;
    font-weight: 600;
    color: #f9fafb;
    letter-spacing: -1px;
    line-height: 1;
}}
.kpi-delta {{
    font-size: 0.72rem;
    margin-top: 8px;
    color: rgba(255,255,255,0.35);
}}
.kpi-delta.up   {{ color: {COR_SUCESSO}; }}
.kpi-delta.down {{ color: {COR_ERRO}; }}

/* ── STATUS BADGE ── */
.badge {{
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 0.3px;
}}

/* ── FORM INPUTS ── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea {{
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    border-radius: 8px !important;
    color: #f9fafb !important;
    font-size: 0.875rem !important;
}}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {{
    border-color: {COR_PRIMARIA} !important;
    box-shadow: 0 0 0 2px rgba(197,143,61,0.12) !important;
    background: rgba(255,255,255,0.05) !important;
}}
.stSelectbox > div > div,
.stMultiSelect > div > div {{
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    border-radius: 8px !important;
}}
.stSlider > div > div > div > div {{
    background: {COR_PRIMARIA} !important;
}}
.stRadio > div > label {{
    background: rgba(255,255,255,0.03) !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 8px !important;
    padding: 10px 16px !important;
    transition: all 0.2s !important;
}}
.stRadio > div > label:hover {{
    border-color: rgba(255,255,255,0.14) !important;
}}
.stRadio > div > label[data-checked="true"] {{
    border-color: {COR_PRIMARIA} !important;
    background: rgba(197,143,61,0.08) !important;
}}

/* ── MAIN CONTENT BUTTONS ── */
.block-container div.stButton > button {{
    background: transparent;
    border: 1px solid rgba(255,255,255,0.1);
    color: #e5e7eb;
    border-radius: 8px;
    padding: 10px 20px;
    font-size: 0.82rem;
    font-weight: 500;
    letter-spacing: 0.3px;
    text-transform: none;
    transition: all 0.2s;
}}
.block-container div.stButton > button:hover {{
    border-color: {COR_PRIMARIA};
    color: {COR_PRIMARIA};
    transform: translateY(-1px);
    box-shadow: 0 4px 16px rgba(197,143,61,0.15);
}}
.block-container div.stButton > button[kind="primary"] {{
    background: linear-gradient(135deg, {COR_PRIMARIA}, #a8793a);
    border: none;
    color: #ffffff;
    font-weight: 600;
    box-shadow: 0 2px 10px rgba(197,143,61,0.25);
}}
.block-container div.stButton > button[kind="primary"]:hover {{
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(197,143,61,0.4);
    color: #ffffff !important;
}}

/* ── DATAFRAME ── */
.stDataFrame {{ border-radius: 10px; overflow: hidden; }}
[data-testid="stDataFrameResizable"] {{ background: #111111 !important; }}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {{
    background: rgba(255,255,255,0.02) !important;
    border-radius: 10px !important;
    padding: 4px !important;
    border: 1px solid rgba(255,255,255,0.06) !important;
    gap: 4px !important;
}}
.stTabs [data-baseweb="tab"] {{
    border-radius: 7px !important;
    color: rgba(255,255,255,0.45) !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    padding: 8px 16px !important;
}}
.stTabs [aria-selected="true"] {{
    background: rgba(197,143,61,0.15) !important;
    color: {COR_PRIMARIA} !important;
}}

/* ── EXPANDER ── */
.streamlit-expanderHeader {{
    background: rgba(255,255,255,0.02) !important;
    border: 1px solid rgba(255,255,255,0.06) !important;
    border-radius: 8px !important;
    color: #e5e7eb !important;
}}

/* ── ALERTS ── */
.pro-alert {{
    padding: 12px 16px;
    border-radius: 8px;
    font-size: 0.82rem;
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 16px;
}}
.pro-alert.info    {{ background: rgba(59,130,246,0.08);  border: 1px solid rgba(59,130,246,0.2);  color: #93c5fd; }}
.pro-alert.success {{ background: rgba(16,185,129,0.08); border: 1px solid rgba(16,185,129,0.2); color: #6ee7b7; }}
.pro-alert.warning {{ background: rgba(245,158,11,0.08); border: 1px solid rgba(245,158,11,0.2); color: #fcd34d; }}
.pro-alert.danger  {{ background: rgba(239,68,68,0.08);  border: 1px solid rgba(239,68,68,0.2);  color: #fca5a5; }}

/* ── DIVIDER ── */
.pro-divider {{
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.08), transparent);
    margin: 28px 0;
}}

/* ── TEAM ROW ── */
.team-row {{
    display: flex;
    align-items: center;
    padding: 14px 18px;
    background: rgba(255,255,255,0.02);
    border: 1px solid rgba(255,255,255,0.04);
    border-radius: 10px;
    margin-bottom: 6px;
    transition: all 0.2s;
    gap: 14px;
}}
.team-row:hover {{
    background: rgba(255,255,255,0.04);
    border-color: rgba(255,255,255,0.08);
}}
.team-avatar {{
    width: 36px; height: 36px;
    border-radius: 8px;
    background: linear-gradient(135deg, rgba(197,143,61,0.25), rgba(197,143,61,0.1));
    display: flex; align-items: center; justify-content: center;
    font-weight: 700; font-size: 0.85rem;
    color: {COR_PRIMARIA};
    flex-shrink: 0;
}}
.team-info {{ flex: 1; }}
.team-name {{ font-weight: 600; color: #e5e7eb; font-size: 0.85rem; }}
.team-sub  {{ font-size: 0.72rem; color: rgba(255,255,255,0.35); margin-top: 2px; }}
.team-cost {{
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.85rem;
    color: {COR_SUCESSO};
    font-weight: 500;
}}

/* ── HIDE STREAMLIT DEFAULTS ── */
#MainMenu {{ visibility: hidden; }}
footer    {{ visibility: hidden; }}
header    {{ visibility: hidden; }}
.stDeployButton {{ display: none; }}
div[data-testid="stToolbar"] {{ display: none; }}

/* ── STEP INDICATOR ── */
.step-bar {{
    display: flex; align-items: center;
    gap: 0; margin-bottom: 36px;
    padding: 0;
}}
.step-item {{
    display: flex; align-items: center; gap: 10px;
    flex: 1;
    padding: 12px 20px;
    border-radius: 8px;
    font-size: 0.78rem;
    font-weight: 500;
    letter-spacing: 0.3px;
    text-transform: uppercase;
    transition: all 0.3s;
}}
.step-item.active {{
    background: rgba(197,143,61,0.12);
    border: 1px solid rgba(197,143,61,0.4);
    color: {COR_PRIMARIA};
}}
.step-item.done {{
    background: rgba(16,185,129,0.06);
    border: 1px solid rgba(16,185,129,0.2);
    color: {COR_SUCESSO};
}}
.step-item.pending {{
    background: rgba(255,255,255,0.02);
    border: 1px solid rgba(255,255,255,0.05);
    color: rgba(255,255,255,0.2);
}}
.step-num {{
    width: 22px; height: 22px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.68rem; font-weight: 700;
    flex-shrink: 0;
}}
.step-item.active  .step-num {{ background: {COR_PRIMARIA}; color: #fff; }}
.step-item.done    .step-num {{ background: {COR_SUCESSO};  color: #fff; }}
.step-item.pending .step-num {{ background: rgba(255,255,255,0.08); color: rgba(255,255,255,0.25); }}
.step-connector {{
    width: 24px; height: 2px;
    background: rgba(255,255,255,0.06);
    flex-shrink: 0;
}}
.step-connector.done {{ background: {COR_SUCESSO}; }}
</style>
""", unsafe_allow_html=True)

# ==================== HELPERS ====================

def fmt(v):
    if isinstance(v, Decimal): v = float(v)
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def fmt_k(v):
    if isinstance(v, Decimal): v = float(v)
    if v >= 1_000_000: return f"R$ {v/1_000_000:.1f}M"
    if v >= 1_000:     return f"R$ {v/1_000:.1f}K"
    return f"R$ {v:.0f}"

def calc_custo(salario, ded_pct, meses, qtd=1):
    s = Decimal(str(salario))
    d = Decimal(str(ded_pct)) / Decimal("100")
    m = Decimal(str(meses))
    q = Decimal(str(qtd))
    mensal_cheio   = s * FATOR_ENCARGOS + BENEFICIOS_MENSAIS
    mensal_dedicado = mensal_cheio * d
    total = mensal_dedicado * m * q
    return float(mensal_cheio), float(mensal_dedicado), float(total)

def calc_hora(salario):
    s = Decimal(str(salario))
    return float((s * FATOR_ENCARGOS + BENEFICIOS_MENSAIS) / HORAS_MES)

def calc_preco(custo, gm, imp, com):
    div = 1 - (gm + imp + com) / 100
    if div <= 0: return None, 0, 0, 0
    p = custo / div
    return p, p * imp / 100, p * com / 100, p * gm / 100

def badge_html(status):
    cor = STATUS_COLORS.get(status, "#6b7280")
    return f'<span class="badge" style="background:rgba(0,0,0,0);border:1px solid {cor};color:{cor};">{status}</span>'

def chart_layout(**kwargs):
    base = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color="#6b7280", size=11),
        margin=dict(t=30, b=20, l=10, r=10),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#9ca3af", size=10)),
        xaxis=dict(gridcolor="rgba(255,255,255,0.04)", zerolinecolor="rgba(255,255,255,0.06)", tickfont=dict(color="#6b7280")),
        yaxis=dict(gridcolor="rgba(255,255,255,0.04)", zerolinecolor="rgba(255,255,255,0.06)", tickfont=dict(color="#6b7280")),
    )
    base.update(kwargs)
    return base

def kpi_card(label, value, delta=None, color=COR_PRIMARIA, mono=True):
    delta_html = ""
    if delta is not None:
        cls = "up" if delta >= 0 else "down"
        icon = "↑" if delta >= 0 else "↓"
        delta_html = f'<div class="kpi-delta {cls}">{icon} {abs(delta):.1f}%</div>'
    font_class = "font-family:'JetBrains Mono',monospace;" if mono else ""
    return f"""
<div class="kpi-wrap" style="--kpi-color:{color};">
  <div class="kpi-label">{label}</div>
  <div class="kpi-value" style="{font_class}">{value}</div>
  {delta_html}
</div>"""

def step_bar(atual):
    steps = [("1","Estratégia"),("2","Equipe"),("3","Custos"),("4","Revisão")]
    html  = '<div class="step-bar">'
    for i,(num,label) in enumerate(steps):
        n = i + 1
        cls = "active" if n == atual else ("done" if n < atual else "pending")
        icon = "✓" if cls == "done" else num
        html += f'<div class="step-item {cls}"><div class="step-num">{icon}</div><span>{label}</span></div>'
        if i < len(steps)-1:
            conn_cls = "done" if n < atual else ""
            html += f'<div class="step-connector {conn_cls}"></div>'
    html += '</div>'
    return html

def section_header(icon, title, subtitle=""):
    sub = f'<p style="color:rgba(255,255,255,0.35);font-size:0.82rem;margin:4px 0 0 0;">{subtitle}</p>' if subtitle else ""
    st.markdown(f"""
<div style="margin-bottom:28px;">
  <div style="display:flex;align-items:center;gap:10px;">
    <span style="font-size:1.3rem;">{icon}</span>
    <div>
      <h1 style="margin:0;">{title}</h1>
      {sub}
    </div>
  </div>
</div>""", unsafe_allow_html=True)


# ==================== PAGE: DASHBOARD ====================

def page_dashboard():
    section_header("\U0001f3e0", "Dashboard", "Vis\u00e3o geral da opera\u00e7\u00e3o de vendas")

    rows = db_get_all()
    df   = pd.DataFrame(rows) if rows else pd.DataFrame()

    if df.empty:
        st.markdown('<div class="pro-alert info">\u2b06 Nenhuma proposta ainda. Crie uma em <strong>Nova Proposta</strong>.</div>', unsafe_allow_html=True)
        return

    df["created_at"] = pd.to_datetime(df["created_at"])

    total     = len(df)
    pipeline  = df[df["status"].isin(["Em An\u00e1lise","Aprovada","Em Execu\u00e7\u00e3o"])]["preco_venda"].sum()
    ticket_m  = df["preco_venda"].mean()
    aprovadas = df[df["status"]=="Aprovada"]
    rejeitadas= df[df["status"]=="Rejeitada"]
    denom_wr  = len(aprovadas) + len(rejeitadas)
    win_rate  = (len(aprovadas)/denom_wr*100) if denom_wr else 0
    mg_media  = df["gm_alvo"].mean()

    c1,c2,c3,c4,c5 = st.columns(5)
    with c1: st.markdown(kpi_card("Total de Propostas", total, mono=False, color=COR_ACENTO), unsafe_allow_html=True)
    with c2: st.markdown(kpi_card("Pipeline Ativo", fmt_k(pipeline), color=COR_PRIMARIA), unsafe_allow_html=True)
    with c3: st.markdown(kpi_card("Ticket M\u00e9dio", fmt_k(ticket_m), color=COR_PRIMARIA), unsafe_allow_html=True)
    with c4: st.markdown(kpi_card("Win Rate", f"{win_rate:.1f}%", mono=False, color=COR_SUCESSO), unsafe_allow_html=True)
    with c5: st.markdown(kpi_card("Margem M\u00e9dia", f"{mg_media:.1f}%", mono=False, color=COR_ALERTA), unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)

    col_a, col_b = st.columns([2, 1])
    with col_a:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Receita por M\u00eas (últimos 12 meses)</span></div>', unsafe_allow_html=True)
        cutoff = pd.Timestamp.now() - pd.DateOffset(months=12)
        df_m   = df[df["created_at"] >= cutoff].copy()
        df_m["mes"] = df_m["created_at"].dt.to_period("M").astype(str)
        por_mes = df_m.groupby("mes")["preco_venda"].sum().reset_index()
        por_mes.columns = ["mes","receita"]
        por_mes = por_mes.sort_values("mes")
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=por_mes["mes"], y=por_mes["receita"],
            marker_color=COR_PRIMARIA, marker_opacity=0.85, name="Receita",
            hovertemplate="<b>%{x}</b><br>%{customdata}<extra></extra>",
            customdata=[fmt_k(v) for v in por_mes["receita"]],
        ))
        fig.update_layout(**chart_layout(height=280, showlegend=False))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with col_b:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Status das Propostas</span></div>', unsafe_allow_html=True)
        status_ct = df["status"].value_counts().reset_index()
        status_ct.columns = ["status","count"]
        fig2 = go.Figure(go.Pie(
            labels=status_ct["status"], values=status_ct["count"],
            hole=0.62,
            marker_colors=[STATUS_COLORS.get(s,"#6b7280") for s in status_ct["status"]],
            textinfo="none",
            hovertemplate="<b>%{label}</b><br>%{value} propostas<extra></extra>",
        ))
        fig2.update_layout(**chart_layout(height=280,
            legend=dict(orientation="v", x=1.05, y=0.5, bgcolor="rgba(0,0,0,0)", font=dict(color="#9ca3af",size=10)),
            annotations=[dict(text=f"<b>{total}</b>", x=0.5, y=0.5, font_size=20, font_color="#f9fafb", showarrow=False)],
        ))
        st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    col_c, col_d = st.columns([1.2, 1])
    with col_c:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Top 8 Clientes por Valor</span></div>', unsafe_allow_html=True)
        top_cli = df.groupby("cliente")["preco_venda"].sum().nlargest(8).reset_index()
        fig3 = go.Figure(go.Bar(
            y=top_cli["cliente"], x=top_cli["preco_venda"], orientation="h",
            marker=dict(color=top_cli["preco_venda"], colorscale=[[0,"#1a1a1a"],[1,COR_PRIMARIA]], showscale=False),
            hovertemplate="<b>%{y}</b><br>%{customdata}<extra></extra>",
            customdata=[fmt_k(v) for v in top_cli["preco_venda"]],
        ))
        fig3.update_layout(**chart_layout(height=280, showlegend=False,
            yaxis=dict(autorange="reversed", gridcolor="rgba(255,255,255,0.04)", tickfont=dict(color="#9ca3af",size=10)),
        ))
        st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with col_d:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Receita por Vertical</span></div>', unsafe_allow_html=True)
        by_serv = df.groupby("tipo_servico")["preco_venda"].sum().reset_index()
        by_serv.columns = ["vertical","valor"]
        fig4 = go.Figure(go.Bar(
            x=by_serv["vertical"], y=by_serv["valor"],
            marker_color=CHART_COLORS[:len(by_serv)],
            hovertemplate="<b>%{x}</b><br>%{customdata}<extra></extra>",
            customdata=[fmt_k(v) for v in by_serv["valor"]],
        ))
        fig4.update_layout(**chart_layout(height=280, showlegend=False,
            xaxis=dict(tickfont=dict(size=9, color="#6b7280")),
        ))
        st.plotly_chart(fig4, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Propostas Recentes</span></div>', unsafe_allow_html=True)
    recentes = df.head(10).copy()
    recentes["Data"]        = recentes["created_at"].dt.strftime("%d/%m/%Y")
    recentes["Valor Total"] = recentes["preco_venda"].apply(fmt_k)
    recentes["Fee/M\u00eas"]     = recentes["fee_mensal"].apply(fmt_k)
    st.dataframe(
        recentes[["Data","cliente","projeto","tipo_servico","Valor Total","Fee/M\u00eas","status"]].rename(
            columns={"cliente":"Cliente","projeto":"Projeto","tipo_servico":"Vertical","status":"Status"}),
        hide_index=True, use_container_width=True, height=330,
    )
    st.markdown('</div>', unsafe_allow_html=True)


# ==================== PAGE: NOVA PROPOSTA ====================

def _init_draft():
    if "draft" not in st.session_state or st.session_state.draft is None:
        st.session_state.draft = {
            "id": None, "fase": 1,
            "cliente": "", "segmento": "", "projeto": "", "descricao": "",
            "tipo_contrato": TIPOS_CONTRATO[0], "tipo_servico": TIPOS_SERVICO[0],
            "meses": 12, "imposto_pct": IMPOSTO_PADRAO_PCT,
            "comissao_nb": 0, "comissao_parc": 0, "gm_alvo": GROSS_MARGIN_ALVO,
            "regua_salarial": "mercado", "status": "Em An\u00e1lise",
            "responsavel": "", "equipe": [], "terceiros": [],
            "extras": {"viagens":0.0,"software":0.0,"infraestrutura":0.0,"outros":0.0},
            "obs": "",
        }


def _fase1():
    d = st.session_state.draft
    c1, c2 = st.columns(2)
    with c1:
        d["cliente"]       = st.text_input("Cliente *", value=d["cliente"], placeholder="Ex: Empresa ABC")
        d["tipo_contrato"] = st.selectbox("Modelo de Contrato", TIPOS_CONTRATO,
                                           index=TIPOS_CONTRATO.index(d["tipo_contrato"]))
        d["meses"]         = st.number_input("Dura\u00e7\u00e3o (meses)", 1, 60, d["meses"])
        d["imposto_pct"]   = st.number_input("Al\u00edquota de Imposto (%)", 0.0, 30.0, d["imposto_pct"], 0.5)
    with c2:
        d["projeto"]      = st.text_input("Nome do Projeto *", value=d["projeto"], placeholder="Ex: Transforma\u00e7\u00e3o Digital")
        d["tipo_servico"] = st.selectbox("Vertical de Servi\u00e7o", TIPOS_SERVICO,
                                          index=TIPOS_SERVICO.index(d["tipo_servico"]))
        d["segmento"]     = st.text_input("Segmento do Cliente", value=d["segmento"], placeholder="Ex: Banco, Varejo...")
        d["responsavel"]  = st.text_input("Respons\u00e1vel Brivia", value=d["responsavel"], placeholder="Nome do Account")

    d["descricao"] = st.text_area("Descri\u00e7\u00e3o do Escopo", value=d["descricao"], height=80,
                                   placeholder="Descreva os servi\u00e7os a serem prestados...")

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        d["comissao_nb"]   = st.selectbox("Comiss\u00e3o New Business (%)", COMISSAO_NB,
                                            index=COMISSAO_NB.index(d["comissao_nb"]))
    with c2:
        d["comissao_parc"] = st.selectbox("Comiss\u00e3o Parceiros (%)", COMISSAO_PARCEIROS,
                                            index=COMISSAO_PARCEIROS.index(d["comissao_parc"]))
    with c3:
        d["gm_alvo"] = st.slider("Meta Gross Margin (%)", 20.0, 70.0, d["gm_alvo"], 1.0)

    total_ded = d["imposto_pct"] + d["comissao_nb"] + d["comissao_parc"] + d["gm_alvo"]
    cls = "warning" if total_ded >= 95 else "info"
    st.markdown(f'<div class="pro-alert {cls}">\u24d8 Total de dedu\u00e7\u00f5es: <strong>{total_ded:.1f}%</strong> (Imposto {d["imposto_pct"]}% + NB {d["comissao_nb"]}% + Parceiros {d["comissao_parc"]}% + GM {d["gm_alvo"]}%)</div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    _, col, _ = st.columns([2,1,2])
    with col:
        pode = d["cliente"].strip() and d["projeto"].strip() and total_ded < 100
        if pode:
            if st.button("Pr\u00f3ximo: Equipe \u2192", type="primary", use_container_width=True):
                d["fase"] = 2; st.rerun()
        else:
            st.button("Preencha os campos obrigat\u00f3rios", disabled=True, use_container_width=True)


def _fase2():
    d     = st.session_state.draft
    meses = d["meses"]
    opcoes_regua = {"M\u00e9dia de Mercado":"mercado","Faixa M\u00ednima":"minima","M\u00e9dia Brivia":"brivia"}
    regua_lbl = st.radio("R\u00e9gua Salarial:", list(opcoes_regua.keys()), horizontal=True,
                          index=list(opcoes_regua.values()).index(d["regua_salarial"]))
    d["regua_salarial"] = opcoes_regua[regua_lbl]

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pro-card-header"><span class="pro-card-title">Squad Builder \u2014 {meses} meses</span></div>', unsafe_allow_html=True)

    ca, cb, cc, cd = st.columns([2.5,1.5,1,1])
    with ca: perfil = st.selectbox("Perfil Profissional", list(BASE_SALARIAL.keys()))
    with cb: nivel  = st.selectbox("Senioridade", list(BASE_SALARIAL[perfil].keys()))
    with cc: qtd    = st.number_input("Qtd", 1, 20, 1)
    with cd: ded    = st.number_input("Ded %", 10, 100, 100, 10)

    regua_id = d["regua_salarial"]
    sal      = BASE_SALARIAL[perfil][nivel].get(regua_id, 0)
    _, custo_ded, custo_tot = calc_custo(sal, ded, meses, qtd)

    p1,p2,p3,p4 = st.columns(4)
    for col_, lbl, val in [
        (p1, "Sal\u00e1rio Base",   fmt(sal)),
        (p2, "Custo/Hora",          fmt(calc_hora(sal))),
        (p3, "Custo Mensal",        fmt(custo_ded/qtd if qtd else custo_ded)),
        (p4, f"Custo Total ({qtd}x)", fmt(custo_tot)),
    ]:
        with col_:
            st.markdown(f'<div class="kpi-wrap" style="--kpi-color:{COR_PRIMARIA};"><div class="kpi-label">{lbl}</div><div class="kpi-value" style="font-size:1.1rem;">{val}</div></div>', unsafe_allow_html=True)

    if st.button("+ Adicionar ao Squad", use_container_width=True):
        d["equipe"].append({"perfil":perfil,"nivel":nivel,"qtd":qtd,"dedicacao":ded,"salario_base":sal,"regua":regua_id})
        st.rerun()

    if d["equipe"]:
        st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
        total_eq = 0; total_hc = 0
        for i, item in enumerate(d["equipe"]):
            _, _, tot = calc_custo(item["salario_base"], item["dedicacao"], meses, item["qtd"])
            total_eq += tot; total_hc += item["qtd"]
            col_row, col_del = st.columns([12,1])
            with col_row:
                st.markdown(f'''<div class="team-row">
                  <div class="team-avatar">{item["perfil"][0]}</div>
                  <div class="team-info">
                    <div class="team-name">{item["qtd"]}\u00d7 {item["perfil"]}</div>
                    <div class="team-sub">{item["nivel"]} \u00b7 {item["dedicacao"]}% dedica\u00e7\u00e3o</div>
                  </div>
                  <div class="team-cost">{fmt(tot)}</div>
                </div>''', unsafe_allow_html=True)
            with col_del:
                if st.button("\u2715", key=f"del_eq_{i}"):
                    d["equipe"].pop(i); st.rerun()

        e1, e2 = st.columns(2)
        with e1: st.markdown(kpi_card("Headcount", total_hc, mono=False, color=COR_ACENTO), unsafe_allow_html=True)
        with e2: st.markdown(kpi_card("Custo Total da Equipe", fmt_k(total_eq)), unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    b1, b2 = st.columns(2)
    with b1:
        if st.button("\u2190 Voltar", use_container_width=True): d["fase"]=1; st.rerun()
    with b2:
        if st.button("Pr\u00f3ximo: Custos \u2192", type="primary", use_container_width=True): d["fase"]=3; st.rerun()


def _fase3():
    d     = st.session_state.draft
    meses = d["meses"]

    st.markdown('<div class="pro-card-header"><span class="pro-card-title">Terceiros e Fornecedores</span></div>', unsafe_allow_html=True)
    tc1, tc2, tc3 = st.columns([3,2,1])
    with tc1: desc_t = st.text_input("Descri\u00e7\u00e3o", placeholder="Ex: Freelancer Design, Infra AWS...")
    with tc2: val_t  = st.number_input("Valor Total (R$)", 0.0, step=500.0, format="%.2f")
    with tc3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("+ Add", use_container_width=True):
            if desc_t and val_t > 0:
                d["terceiros"].append({"desc":desc_t,"valor":val_t}); st.rerun()

    if d["terceiros"]:
        total_t = 0
        for i, t in enumerate(d["terceiros"]):
            total_t += t["valor"]
            r1, r2 = st.columns([12,1])
            with r1:
                st.markdown(f'<div class="team-row"><div class="team-avatar">T</div><div class="team-info"><div class="team-name">{t["desc"]}</div><div class="team-sub">Terceirizado \u00b7 valor \u00fanico</div></div><div class="team-cost">{fmt(t["valor"])}</div></div>', unsafe_allow_html=True)
            with r2:
                if st.button("\u2715", key=f"del_t_{i}"): d["terceiros"].pop(i); st.rerun()
        st.markdown(f'<div class="pro-alert info">Total terceiros: <strong>{fmt(total_t)}</strong></div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="pro-card-header"><span class="pro-card-title">OPEX Mensal (\u00d7{meses} meses)</span></div>', unsafe_allow_html=True)
    oc1, oc2 = st.columns(2)
    with oc1:
        d["extras"]["viagens"]        = st.number_input("Viagens e Deslocamentos",  0.0, value=d["extras"]["viagens"],        step=100.0)
        d["extras"]["software"]       = st.number_input("Software e Licen\u00e7as",        0.0, value=d["extras"]["software"],       step=100.0)
    with oc2:
        d["extras"]["infraestrutura"] = st.number_input("Infraestrutura Cloud",     0.0, value=d["extras"]["infraestrutura"], step=100.0)
        d["extras"]["outros"]         = st.number_input("Outros",                   0.0, value=d["extras"]["outros"],         step=100.0)

    opex_m = sum(d["extras"].values())
    ox1, ox2 = st.columns(2)
    with ox1: st.markdown(kpi_card("OPEX Mensal",             fmt_k(opex_m),        color=COR_ALERTA), unsafe_allow_html=True)
    with ox2: st.markdown(kpi_card(f"OPEX Total ({meses}m)", fmt_k(opex_m*meses),  color=COR_ALERTA), unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    d["obs"] = st.text_area("Observa\u00e7\u00f5es / Premissas", value=d["obs"], height=80,
                             placeholder="Riscos, premissas, condi\u00e7\u00f5es especiais...")

    b1, b2 = st.columns(2)
    with b1:
        if st.button("\u2190 Voltar", use_container_width=True): d["fase"]=2; st.rerun()
    with b2:
        if st.button("Revisar e Finalizar \u2192", type="primary", use_container_width=True): d["fase"]=4; st.rerun()


def _fase4():
    d     = st.session_state.draft
    meses = d["meses"]

    custo_eq  = sum(calc_custo(e["salario_base"],e["dedicacao"],meses,e["qtd"])[2] for e in d["equipe"])
    custo_te  = sum(t["valor"] for t in d["terceiros"])
    custo_ex  = sum(d["extras"].values()) * meses
    custo_tot = custo_eq + custo_te + custo_ex
    com_tot   = d["comissao_nb"] + d["comissao_parc"]
    preco, v_imp, v_com, v_gm = calc_preco(custo_tot, d["gm_alvo"], d["imposto_pct"], com_tot)

    if preco is None:
        st.markdown('<div class="pro-alert danger">\u26a0 Soma das dedu\u00e7\u00f5es \u2265 100%. Volte e ajuste as margens.</div>', unsafe_allow_html=True)
        if st.button("\u2190 Voltar \u00e0 Estrat\u00e9gia"): d["fase"]=1; st.rerun()
        return

    fee    = preco / meses
    markup = ((preco / custo_tot) - 1) * 100 if custo_tot else 0

    st.markdown('<div class="pro-card card-accent-left"><div class="pro-card-header"><span class="pro-card-title">Resultado da Precifica\u00e7\u00e3o</span></div>', unsafe_allow_html=True)
    k1,k2,k3,k4 = st.columns(4)
    with k1: st.markdown(kpi_card("Pre\u00e7o de Venda", fmt(preco)), unsafe_allow_html=True)
    with k2: st.markdown(kpi_card("Fee Mensal", fmt(fee)), unsafe_allow_html=True)
    with k3: st.markdown(kpi_card("Margem Bruta", fmt(v_gm), color=COR_SUCESSO), unsafe_allow_html=True)
    with k4: st.markdown(kpi_card("Markup", f"{markup:.1f}%", mono=False, color=COR_ALERTA), unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    co1, co2 = st.columns([1,1.4])
    with co1:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Composi\u00e7\u00e3o do Pre\u00e7o</span></div>', unsafe_allow_html=True)
        fig = go.Figure(go.Pie(
            labels=["Custo Base","Impostos","Comiss\u00f5es","Margem Bruta"],
            values=[custo_tot, v_imp, v_com, v_gm],
            hole=0.62,
            marker_colors=["#374151","#4b5563","#6b7280",COR_PRIMARIA],
            textinfo="percent", textfont_size=11, textfont_color="white",
        ))
        fig.update_layout(**chart_layout(height=260, showlegend=True,
            legend=dict(orientation="h",y=-0.25,x=0.5,xanchor="center"),
            annotations=[dict(text=f"<b>{fmt_k(preco)}</b>",x=0.5,y=0.5,font_size=14,font_color="#f9fafb",showarrow=False)],
        ))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})
        st.markdown('</div>', unsafe_allow_html=True)

    with co2:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Detalhamento Financeiro</span></div>', unsafe_allow_html=True)
        df_det = pd.DataFrame({
            "Categoria":  ["Equipe CLT","Terceiros","OPEX/Extras","= CUSTO BASE","Impostos","Comiss\u00f5es","= MARGEM BRUTA"],
            "Valor":      [fmt(custo_eq),fmt(custo_te),fmt(custo_ex),fmt(custo_tot),fmt(v_imp),fmt(v_com),fmt(v_gm)],
            "% Pre\u00e7o": [f"{custo_eq/preco*100:.1f}%",f"{custo_te/preco*100:.1f}%",
                           f"{custo_ex/preco*100:.1f}%",f"{custo_tot/preco*100:.1f}%",
                           f"{d['imposto_pct']:.1f}%",f"{com_tot:.1f}%",f"{d['gm_alvo']:.1f}%"],
        })
        st.dataframe(df_det, hide_index=True, use_container_width=True, height=260)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    s1, s2, s3 = st.columns(3)
    with s1:
        st.markdown(f"**Cliente:** {d['cliente']}\n\n**Projeto:** {d['projeto']}\n\n**Contrato:** {d['tipo_contrato']}")
    with s2:
        st.markdown(f"**Vertical:** {d['tipo_servico']}\n\n**Dura\u00e7\u00e3o:** {meses} meses\n\n**Oferta:** {MAPEAMENTO_OFERTAS.get(d['tipo_servico'],'-')}")
    with s3:
        hc = sum(e["qtd"] for e in d["equipe"])
        st.markdown(f"**Headcount:** {hc}\n\n**Respons\u00e1vel:** {d['responsavel']}\n\n**GM Alvo:** {d['gm_alvo']:.0f}%")

    st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
    _, col_status, _ = st.columns([1,2,1])
    with col_status:
        d["status"] = st.selectbox("Status inicial da proposta", STATUS_LIST,
                                    index=STATUS_LIST.index(d["status"]))

    b1, b2, b3 = st.columns(3)
    with b1:
        if st.button("\u2190 Voltar", use_container_width=True): d["fase"]=3; st.rerun()
    with b2:
        if st.button("Reiniciar", use_container_width=True):
            st.session_state.draft = None; st.rerun()
    with b3:
        if st.button("Salvar Proposta", type="primary", use_container_width=True):
            hc = sum(e["qtd"] for e in d["equipe"])
            payload = {
                "id": d["id"], "status": d["status"],
                "cliente": d["cliente"], "segmento": d.get("segmento",""),
                "projeto": d["projeto"], "descricao": d.get("descricao",""),
                "tipo_contrato": d["tipo_contrato"], "tipo_servico": d["tipo_servico"],
                "meses": meses, "imposto_pct": d["imposto_pct"],
                "comissao_nb": d["comissao_nb"], "comissao_parc": d["comissao_parc"],
                "gm_alvo": d["gm_alvo"], "regua_salarial": d["regua_salarial"],
                "custo_equipe": custo_eq, "custo_terceiros": custo_te,
                "custo_extras": custo_ex, "custo_total": custo_tot,
                "preco_venda": preco, "fee_mensal": fee,
                "v_impostos": v_imp, "v_comissoes": v_com,
                "margem_bruta": v_gm, "markup_pct": markup,
                "headcount": hc, "responsavel": d.get("responsavel",""),
                "obs": d.get("obs",""),
            }
            db_save(payload, d["equipe"], d["terceiros"])
            st.session_state.draft = None
            st.session_state.pagina = "propostas"
            st.success("Proposta salva com sucesso!")
            st.rerun()


def page_nova_proposta():
    _init_draft()
    d = st.session_state.draft
    section_header("\u2795", "Nova Proposta", "Wizard de precifica\u00e7\u00e3o em 4 etapas")
    st.markdown(step_bar(d["fase"]), unsafe_allow_html=True)
    if d["fase"] == 1: _fase1()
    elif d["fase"] == 2: _fase2()
    elif d["fase"] == 3: _fase3()
    elif d["fase"] == 4: _fase4()


# ==================== PAGE: PROPOSTAS ====================

def page_propostas():
    section_header("\U0001f4cb", "Propostas", "Hist\u00f3rico e gest\u00e3o de todas as propostas")

    with st.expander("\U0001f50d Filtros", expanded=False):
        fc1,fc2,fc3,fc4 = st.columns(4)
        with fc1: f_cli    = st.text_input("Cliente")
        with fc2: f_status = st.selectbox("Status", ["Todos"]+STATUS_LIST)
        with fc3: f_serv   = st.selectbox("Vertical", ["Todas"]+TIPOS_SERVICO)
        with fc4: f_cont   = st.selectbox("Contrato",  ["Todos"]+TIPOS_CONTRATO)

    filtros = {}
    if f_cli:               filtros["cliente"]       = f_cli
    if f_status != "Todos": filtros["status"]        = f_status
    if f_serv   != "Todas": filtros["tipo_servico"]  = f_serv
    if f_cont   != "Todos": filtros["tipo_contrato"] = f_cont

    rows = db_get_all(filtros)
    if not rows:
        st.markdown('<div class="pro-alert info">Nenhuma proposta encontrada.</div>', unsafe_allow_html=True)
        return

    st.markdown(f'<p style="color:rgba(255,255,255,0.35);font-size:0.78rem;margin-bottom:12px;">{len(rows)} proposta(s) encontrada(s)</p>', unsafe_allow_html=True)

    for row in rows:
        cor_status = STATUS_COLORS.get(row["status"],"#6b7280")
        with st.expander(f"#{row['id']:04d} · {row['cliente']} — {row['projeto']}  [{row['status']}]"):
            dc1,dc2,dc3 = st.columns(3)
            with dc1:
                st.markdown(f"**Cliente:** {row['cliente']}\n\n**Segmento:** {row.get('segmento','-')}\n\n**Respons\u00e1vel:** {row.get('responsavel','-')}")
            with dc2:
                st.markdown(f"**Contrato:** {row['tipo_contrato']}\n\n**Vertical:** {row['tipo_servico']}\n\n**Dura\u00e7\u00e3o:** {row['meses']} meses")
            with dc3:
                st.markdown(f"**Criado em:** {row['created_at'][:10]}\n\n**Headcount:** {row['headcount']}\n\n**GM Alvo:** {row['gm_alvo']:.0f}%")

            k1,k2,k3,k4 = st.columns(4)
            with k1: st.markdown(kpi_card("Pre\u00e7o de Venda", fmt_k(row["preco_venda"])), unsafe_allow_html=True)
            with k2: st.markdown(kpi_card("Fee Mensal",          fmt_k(row["fee_mensal"])), unsafe_allow_html=True)
            with k3: st.markdown(kpi_card("Custo Total",         fmt_k(row["custo_total"])), unsafe_allow_html=True)
            with k4: st.markdown(kpi_card("Markup", f"{row['markup_pct']:.1f}%", mono=False, color=COR_ALERTA), unsafe_allow_html=True)

            _, eq, _ = db_get_one(row["id"])
            if eq:
                st.markdown("**Equipe:**")
                for e in eq:
                    st.markdown(f"- {e['qtd']}\u00d7 {e['perfil']} ({e['nivel']}) \u00b7 {e['dedicacao']}% ded.")

            st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
            ac1, ac2, ac3 = st.columns(3)
            with ac1:
                novo_st = st.selectbox("Alterar Status", STATUS_LIST,
                                        index=STATUS_LIST.index(row["status"]),
                                        key=f"st_{row['id']}")
                if novo_st != row["status"]:
                    if st.button("Salvar Status", key=f"sv_{row['id']}", type="primary"):
                        db_update_status(row["id"], novo_st); st.rerun()
            with ac3:
                if st.button("\U0001f5d1 Excluir", key=f"del_{row['id']}"):
                    db_delete(row["id"]); st.rerun()


# ==================== PAGE: ANALYTICS ====================

def page_analytics():
    section_header("\U0001f4ca", "Analytics", "An\u00e1lise profunda do pipeline e desempenho")

    rows = db_get_all()
    if not rows:
        st.markdown('<div class="pro-alert info">Sem dados para analytics.</div>', unsafe_allow_html=True)
        return

    df = pd.DataFrame(rows)
    df["created_at"] = pd.to_datetime(df["created_at"])
    df["mes"]  = df["created_at"].dt.to_period("M").astype(str)
    df["trim"] = df["created_at"].dt.to_period("Q").astype(str)

    tabs = st.tabs(["\U0001f4c8 Receita","\U0001f3af Funil","\U0001f4d0 Margens","\U0001f5c2 Verticais","\U0001f465 Equipe","\U0001f50d Scatter"])

    with tabs[0]:
        gran = st.radio("Granularidade:", ["Mensal","Trimestral"], horizontal=True)
        grp  = "mes" if gran == "Mensal" else "trim"
        rev  = df.groupby(grp).agg(receita=("preco_venda","sum"), qtd=("id","count")).reset_index().sort_values(grp)
        fig  = go.Figure()
        fig.add_trace(go.Bar(x=rev[grp], y=rev["receita"], name="Receita",
            marker_color=COR_PRIMARIA, marker_opacity=0.8,
            hovertemplate="<b>%{x}</b><br>%{customdata[0]} · %{customdata[1]} propostas<extra></extra>",
            customdata=list(zip([fmt_k(v) for v in rev["receita"]], rev["qtd"])),
        ))
        fig.add_trace(go.Scatter(x=rev[grp], y=rev["receita"].rolling(3,min_periods=1).mean(),
            name="M\u00e9dia M\u00f3vel 3p", mode="lines", line=dict(color=COR_ACENTO, width=2, dash="dot"),
        ))
        fig.update_layout(**chart_layout(height=340, showlegend=True,
            legend=dict(orientation="h",y=1.05,x=0,xanchor="left")))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})

        by_cont = df.groupby(["mes","tipo_contrato"])["preco_venda"].sum().reset_index()
        fig2 = go.Figure()
        for i, cont in enumerate(TIPOS_CONTRATO):
            sub = by_cont[by_cont["tipo_contrato"]==cont].sort_values("mes")
            fig2.add_trace(go.Bar(x=sub["mes"], y=sub["preco_venda"],
                name=cont.split(" ")[0], marker_color=CHART_COLORS[i % len(CHART_COLORS)]))
        fig2.update_layout(**chart_layout(height=280, barmode="stack", showlegend=True,
            title=dict(text="Receita por Tipo de Contrato",font=dict(color="#9ca3af",size=12))))
        st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar":False})

    with tabs[1]:
        ordem_funil = ["Rascunho","Em An\u00e1lise","Aprovada","Em Execu\u00e7\u00e3o","Conclу\u00edda"]
        ordem_funil = ["Rascunho","Em Análise","Aprovada","Em Execução","Concluída"]
        funil_ct = df[df["status"].isin(ordem_funil)]["status"].value_counts()
        funil_df = pd.DataFrame({"status":ordem_funil,"count":[funil_ct.get(s,0) for s in ordem_funil]})

        f1, f2 = st.columns(2)
        with f1:
            fig_f = go.Figure(go.Funnel(
                y=funil_df["status"], x=funil_df["count"],
                textinfo="value+percent initial",
                marker=dict(color=[STATUS_COLORS.get(s,"#6b7280") for s in ordem_funil]),
            ))
            fig_f.update_layout(**chart_layout(height=360, showlegend=False,
                title=dict(text="Volume de Propostas por Est\u00e1gio",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_f, use_container_width=True, config={"displayModeBar":False})
        with f2:
            ap = len(df[df["status"]=="Aprovada"]); re = len(df[df["status"]=="Rejeitada"])
            wr = ap/(ap+re)*100 if (ap+re) else 0
            fig_g = go.Figure(go.Indicator(
                mode="gauge+number+delta", value=wr,
                delta={"reference":50,"valueformat":".1f"},
                number={"suffix":"%","font":{"size":36,"color":"#f9fafb"}},
                gauge=dict(
                    axis=dict(range=[0,100],tickfont=dict(color="#6b7280")),
                    bar=dict(color=COR_SUCESSO), bgcolor="rgba(255,255,255,0.03)",
                    steps=[dict(range=[0,40],color="rgba(239,68,68,0.1)"),
                           dict(range=[40,70],color="rgba(245,158,11,0.1)"),
                           dict(range=[70,100],color="rgba(16,185,129,0.1)")],
                    threshold=dict(line=dict(color=COR_PRIMARIA,width=2),value=50),
                ),
                title=dict(text="Win Rate Global",font=dict(color="#9ca3af",size=14)),
            ))
            fig_g.update_layout(**chart_layout(height=360))
            st.plotly_chart(fig_g, use_container_width=True, config={"displayModeBar":False})

        wr_data = []
        for serv in TIPOS_SERVICO:
            sub = df[df["tipo_servico"]==serv]
            a = len(sub[sub["status"]=="Aprovada"]); r = len(sub[sub["status"]=="Rejeitada"])
            wr_data.append({"vertical":serv,"win_rate":a/(a+r)*100 if (a+r) else 0})
        wr_df = pd.DataFrame(wr_data)
        fig_wr = go.Figure(go.Bar(
            x=wr_df["vertical"], y=wr_df["win_rate"],
            marker_color=[COR_SUCESSO if v>=50 else COR_ALERTA for v in wr_df["win_rate"]],
            text=[f"{v:.1f}%" for v in wr_df["win_rate"]], textposition="outside", textfont_color="#9ca3af",
        ))
        fig_wr.update_layout(**chart_layout(height=260, showlegend=False,
            title=dict(text="Win Rate por Vertical",font=dict(color="#9ca3af",size=12)), yaxis=dict(range=[0,110])))
        st.plotly_chart(fig_wr, use_container_width=True, config={"displayModeBar":False})

    with tabs[2]:
        m1, m2 = st.columns(2)
        with m1:
            fig_h = go.Figure(go.Histogram(x=df["gm_alvo"], nbinsx=15,
                marker_color=COR_PRIMARIA, marker_opacity=0.8))
            fig_h.update_layout(**chart_layout(height=300, showlegend=False,
                title=dict(text="Distribui\u00e7\u00e3o de Gross Margin",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_h, use_container_width=True, config={"displayModeBar":False})
        with m2:
            fig_b = go.Figure(go.Box(y=df["markup_pct"], boxpoints="outliers",
                marker_color=COR_PRIMARIA, line_color=COR_ACENTO,
                fillcolor="rgba(197,143,61,0.1)", name="Markup %"))
            fig_b.update_layout(**chart_layout(height=300, showlegend=False,
                title=dict(text="Distribui\u00e7\u00e3o de Markup",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_b, use_container_width=True, config={"displayModeBar":False})

        fig_v = go.Figure()
        for i, cont in enumerate(TIPOS_CONTRATO):
            sub = df[df["tipo_contrato"]==cont]["fee_mensal"]
            if len(sub) >= 3:
                rgb = tuple(int(CHART_COLORS[i % len(CHART_COLORS)].lstrip("#")[j*2:(j+1)*2],16) for j in range(3))
                fig_v.add_trace(go.Violin(y=sub, name=cont.split(" ")[0], box_visible=True, meanline_visible=True,
                    fillcolor=f"rgba({rgb[0]},{rgb[1]},{rgb[2]},0.15)",
                    line_color=CHART_COLORS[i % len(CHART_COLORS)]))
        fig_v.update_layout(**chart_layout(height=300, showlegend=True,
            title=dict(text="Fee Mensal por Tipo de Contrato",font=dict(color="#9ca3af",size=12))))
        st.plotly_chart(fig_v, use_container_width=True, config={"displayModeBar":False})

    with tabs[3]:
        by_serv = df.groupby("tipo_servico").agg(
            receita=("preco_venda","sum"), count=("id","count"),
            ticket=("preco_venda","mean"), gm=("gm_alvo","mean"),
        ).reset_index()
        fig_tm = px.treemap(by_serv, path=["tipo_servico"], values="receita",
            color="gm", color_continuous_scale=[[0,"#1a1a1a"],[0.5,COR_ALERTA],[1,COR_PRIMARIA]],
            custom_data=["count","ticket","gm"])
        fig_tm.update_traces(
            hovertemplate="<b>%{label}</b><br>Receita: %{value:,.0f}<br>Propostas: %{customdata[0]}<br>GM: %{customdata[2]:.1f}%<extra></extra>",
            textfont_size=13)
        fig_tm.update_layout(**chart_layout(height=360, coloraxis_showscale=False,
            title=dict(text="Treemap de Receita por Vertical (cor = GM m\u00e9dio)",font=dict(color="#9ca3af",size=12))))
        st.plotly_chart(fig_tm, use_container_width=True, config={"displayModeBar":False})

        t1,t2,t3 = st.columns(3)
        with t1:
            fig_tk = go.Figure(go.Bar(y=by_serv["tipo_servico"], x=by_serv["ticket"], orientation="h",
                marker_color=CHART_COLORS[:len(by_serv)],
                customdata=[fmt_k(v) for v in by_serv["ticket"]],
                hovertemplate="<b>%{y}</b><br>%{customdata}<extra></extra>"))
            fig_tk.update_layout(**chart_layout(height=260,showlegend=False,
                title=dict(text="Ticket M\u00e9dio",font=dict(color="#9ca3af",size=12)),yaxis=dict(autorange="reversed")))
            st.plotly_chart(fig_tk, use_container_width=True, config={"displayModeBar":False})
        with t2:
            fig_ct = go.Figure(go.Bar(y=by_serv["tipo_servico"], x=by_serv["count"], orientation="h",
                marker_color=COR_ACENTO, marker_opacity=0.75))
            fig_ct.update_layout(**chart_layout(height=260,showlegend=False,
                title=dict(text="Volume de Propostas",font=dict(color="#9ca3af",size=12)),yaxis=dict(autorange="reversed")))
            st.plotly_chart(fig_ct, use_container_width=True, config={"displayModeBar":False})
        with t3:
            fig_gm2 = go.Figure(go.Bar(y=by_serv["tipo_servico"], x=by_serv["gm"], orientation="h",
                marker_color=COR_SUCESSO, marker_opacity=0.75,
                text=[f"{v:.1f}%" for v in by_serv["gm"]], textposition="outside", textfont_color="#9ca3af"))
            fig_gm2.update_layout(**chart_layout(height=260,showlegend=False,
                title=dict(text="GM M\u00e9dio",font=dict(color="#9ca3af",size=12)),yaxis=dict(autorange="reversed")))
            st.plotly_chart(fig_gm2, use_container_width=True, config={"displayModeBar":False})

    with tabs[4]:
        e1, e2 = st.columns(2)
        with e1:
            fig_hc = go.Figure(go.Histogram(x=df["headcount"], nbinsx=10,
                marker_color=COR_ACENTO, marker_opacity=0.8))
            fig_hc.update_layout(**chart_layout(height=280,showlegend=False,
                title=dict(text="Distribui\u00e7\u00e3o de Headcount",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_hc, use_container_width=True, config={"displayModeBar":False})
        with e2:
            df["pct_eq"] = (df["custo_equipe"]/df["custo_total"].replace(0,1)*100).fillna(0)
            fig_peq = go.Figure(go.Histogram(x=df["pct_eq"], nbinsx=12,
                marker_color=COR_ALERTA, marker_opacity=0.8))
            fig_peq.update_layout(**chart_layout(height=280,showlegend=False,
                title=dict(text="% Custo Equipe / Custo Total",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_peq, use_container_width=True, config={"displayModeBar":False})

        dur_df = df.groupby("tipo_servico")["meses"].mean().reset_index()
        fig_dur = go.Figure(go.Bar(x=dur_df["tipo_servico"], y=dur_df["meses"],
            marker_color=CHART_COLORS[:len(dur_df)],
            text=[f"{v:.1f}m" for v in dur_df["meses"]], textposition="outside", textfont_color="#9ca3af"))
        fig_dur.update_layout(**chart_layout(height=260,showlegend=False,
            title=dict(text="Dura\u00e7\u00e3o M\u00e9dia por Vertical (meses)",font=dict(color="#9ca3af",size=12))))
        st.plotly_chart(fig_dur, use_container_width=True, config={"displayModeBar":False})

    with tabs[5]:
        sc1, sc2 = st.columns(2)
        with sc1:
            fig_sc1 = px.scatter(df, x="headcount", y="markup_pct",
                color="tipo_servico", size="preco_venda", size_max=40,
                color_discrete_sequence=CHART_COLORS,
                hover_data={"cliente":True,"projeto":True,"preco_venda":False},
                labels={"headcount":"Headcount","markup_pct":"Markup (%)","tipo_servico":"Vertical"})
            fig_sc1.update_layout(**chart_layout(height=360,
                title=dict(text="Headcount vs Markup (tamanho = valor)",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_sc1, use_container_width=True, config={"displayModeBar":False})
        with sc2:
            fig_sc2 = px.scatter(df, x="meses", y="preco_venda",
                color="tipo_contrato", size="headcount", size_max=35,
                color_discrete_sequence=CHART_COLORS,
                hover_data={"cliente":True,"status":True},
                labels={"meses":"Dura\u00e7\u00e3o (meses)","preco_venda":"Valor Total (R$)","tipo_contrato":"Contrato"})
            fig_sc2.update_layout(**chart_layout(height=360,
                title=dict(text="Dura\u00e7\u00e3o vs Valor Total",font=dict(color="#9ca3af",size=12))))
            st.plotly_chart(fig_sc2, use_container_width=True, config={"displayModeBar":False})


# ==================== PAGE: IMPORTAR ====================

def _gerar_template_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importar Propostas"
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    gold  = PatternFill("solid", fgColor="C58F3D")
    dark  = PatternFill("solid", fgColor="111111")
    med   = PatternFill("solid", fgColor="1C1C1C")
    wbold = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    wreg  = Font(color="CCCCCC", name="Calibri", size=10)
    ctr   = Alignment(horizontal="center", vertical="center")
    thin  = Border(
        left=Side(style="thin",color="2D2D2D"), right=Side(style="thin",color="2D2D2D"),
        top=Side(style="thin",color="2D2D2D"),  bottom=Side(style="thin",color="2D2D2D"),
    )
    headers = ["Cliente*","Segmento","Projeto*","Descri\u00e7\u00e3o",
               "Tipo Contrato*","Tipo Servi\u00e7o*","Meses*",
               "Imposto %","Comiss\u00e3o NB %","Comiss\u00e3o Parceiros %","GM Alvo %",
               "R\u00e9gua Salarial","Respons\u00e1vel","Status","Observa\u00e7\u00f5es"]
    widths  = [22,18,28,30,25,22,8,10,14,18,10,15,20,14,25]
    ws.row_dimensions[1].height = 30
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=1,column=ci,value=h)
        cell.fill=gold; cell.font=wbold; cell.alignment=ctr; cell.border=thin
        ws.column_dimensions[get_column_letter(ci)].width=w

    exemplos=[
        ["Empresa XYZ","Varejo","Portal B2C","Desenvolvimento do portal","Fee Mensal (Recorrente)","Tecnologia / Dev",12,14.25,5,0,45,"mercado","Ana Paula Ferreira","Em An\u00e1lise","Projeto Q3"],
        ["Banco ABC","Banco","Data Lake","Data lake corporativo","Projeto (Escopo Fechado)","Dados / Analytics",6,13.5,7,5,40,"brivia","Carlos Lima","Rascunho",""],
    ]
    for ri,ex in enumerate(exemplos,2):
        ws.row_dimensions[ri].height=20
        for ci,v in enumerate(ex,1):
            cell=ws.cell(row=ri,column=ci,value=v)
            cell.fill=med if ri%2==0 else dark; cell.font=wreg
            cell.alignment=Alignment(vertical="center"); cell.border=thin

    ws2=wb.create_sheet("Refer\u00eancias")
    ws2["A1"]="Tipos de Contrato"; ws2["B1"]="Tipos de Servi\u00e7o"; ws2["C1"]="Status"; ws2["D1"]="R\u00e9gua"
    for i,v in enumerate(TIPOS_CONTRATO,2): ws2[f"A{i}"]=v
    for i,v in enumerate(TIPOS_SERVICO,2):  ws2[f"B{i}"]=v
    for i,v in enumerate(STATUS_LIST,2):    ws2[f"C{i}"]=v
    for i,v in enumerate(["mercado","minima","brivia"],2): ws2[f"D{i}"]=v

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def page_importar():
    section_header("\U0001f4e4", "Importar Dados", "Importe propostas via planilha Excel")

    if not EXCEL_OK:
        st.markdown('<div class="pro-alert danger">\u26a0 Instale <code>openpyxl</code>: <code>pip install openpyxl</code></div>', unsafe_allow_html=True)
        return

    st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">1 \u00b7 Baixe o Template</span></div>', unsafe_allow_html=True)
    st.markdown('<p>Baixe o template Excel formatado, preencha e fa\u00e7a o upload abaixo.</p>', unsafe_allow_html=True)
    st.download_button("\u2b07 Baixar Template Excel", data=_gerar_template_excel(),
        file_name="brivia_pricing_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">2 \u00b7 Upload do Arquivo</span></div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Selecione o arquivo .xlsx preenchido", type=["xlsx"])

    if uploaded:
        try:
            df_i = pd.read_excel(uploaded, sheet_name="Importar Propostas", header=0)
            df_i.columns = df_i.columns.str.replace("*","",regex=False).str.strip()
            df_i = df_i.dropna(subset=["Cliente","Projeto"])

            st.markdown(f'<div class="pro-alert success">\u2713 {len(df_i)} linha(s) detectada(s).</div>', unsafe_allow_html=True)
            st.dataframe(df_i, hide_index=True, use_container_width=True, height=200)

            erros=[]
            for i,row in df_i.iterrows():
                if str(row.get("Tipo Contrato","")) not in TIPOS_CONTRATO:
                    erros.append(f"Linha {i+2}: Tipo Contrato inv\u00e1lido")
                if str(row.get("Tipo Servi\u00e7o","")) not in TIPOS_SERVICO:
                    erros.append(f"Linha {i+2}: Tipo Servi\u00e7o inv\u00e1lido")

            for e in erros:
                st.markdown(f'<div class="pro-alert danger">\u2715 {e}</div>', unsafe_allow_html=True)

            if not erros:
                if st.button("\u2705 Importar para o Banco", type="primary"):
                    count=0
                    for _,row in df_i.iterrows():
                        payload={
                            "id":None,"status":str(row.get("Status","Rascunho") or "Rascunho"),
                            "cliente":str(row["Cliente"]),"segmento":str(row.get("Segmento","") or ""),
                            "projeto":str(row["Projeto"]),"descricao":str(row.get("Descri\u00e7\u00e3o","") or ""),
                            "tipo_contrato":str(row.get("Tipo Contrato",TIPOS_CONTRATO[0])),
                            "tipo_servico":str(row.get("Tipo Servi\u00e7o",TIPOS_SERVICO[0])),
                            "meses":int(row.get("Meses",12) or 12),
                            "imposto_pct":float(row.get("Imposto %",IMPOSTO_PADRAO_PCT) or IMPOSTO_PADRAO_PCT),
                            "comissao_nb":float(row.get("Comiss\u00e3o NB %",0) or 0),
                            "comissao_parc":float(row.get("Comiss\u00e3o Parceiros %",0) or 0),
                            "gm_alvo":float(row.get("GM Alvo %",GROSS_MARGIN_ALVO) or GROSS_MARGIN_ALVO),
                            "regua_salarial":str(row.get("R\u00e9gua Salarial","mercado") or "mercado"),
                            "custo_equipe":0,"custo_terceiros":0,"custo_extras":0,"custo_total":0,
                            "preco_venda":0,"fee_mensal":0,"v_impostos":0,"v_comissoes":0,
                            "margem_bruta":0,"markup_pct":0,"headcount":0,
                            "responsavel":str(row.get("Respons\u00e1vel","") or ""),
                            "obs":str(row.get("Observa\u00e7\u00f5es","") or ""),
                        }
                        db_save(payload,[],[])
                        count+=1
                    st.markdown(f'<div class="pro-alert success">\u2713 {count} proposta(s) importada(s)!</div>', unsafe_allow_html=True)
        except Exception as ex:
            st.markdown(f'<div class="pro-alert danger">Erro ao ler arquivo: {ex}</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ==================== PAGE: CONFIGURAÇÕES ====================

def page_config():
    section_header("\u2699\ufe0f", "Configura\u00e7\u00f5es", "Par\u00e2metros do sistema e informa\u00e7\u00f5es do banco")

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Par\u00e2metros Padr\u00e3o</span></div>', unsafe_allow_html=True)
        if "cfg_imposto" not in st.session_state: st.session_state.cfg_imposto    = IMPOSTO_PADRAO_PCT
        if "cfg_gm"      not in st.session_state: st.session_state.cfg_gm         = GROSS_MARGIN_ALVO
        if "cfg_regua"   not in st.session_state: st.session_state.cfg_regua      = "mercado"
        if "cfg_benef"   not in st.session_state: st.session_state.cfg_benef      = float(BENEFICIOS_MENSAIS)

        st.session_state.cfg_imposto = st.number_input("Al\u00edquota de Imposto Padr\u00e3o (%)", 0.0, 30.0, st.session_state.cfg_imposto, 0.25)
        st.session_state.cfg_gm      = st.slider("Gross Margin Alvo Padr\u00e3o (%)", 20.0, 70.0, st.session_state.cfg_gm, 1.0)
        st.session_state.cfg_regua   = st.radio("R\u00e9gua Salarial Padr\u00e3o", ["mercado","minima","brivia"], horizontal=True,
                                                  index=["mercado","minima","brivia"].index(st.session_state.cfg_regua))
        st.session_state.cfg_benef   = st.number_input("Benef\u00edcios Mensais/CLT (R$)", 0.0, 5000.0, st.session_state.cfg_benef, 50.0,
                                                         help="VR + VA + Plano de Sa\u00fade m\u00e9dio")
        st.markdown('<div class="pro-alert info">\u24d8 Estes par\u00e2metros afetam novas propostas criadas nesta sess\u00e3o.</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with cc2:
        st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Informa\u00e7\u00f5es do Banco</span></div>', unsafe_allow_html=True)
        total_p = db_count()
        db_size = DB_PATH.stat().st_size / 1024 if DB_PATH.exists() else 0
        rows_inf = db_get_all()
        df_inf   = pd.DataFrame(rows_inf) if rows_inf else pd.DataFrame()

        st.markdown(kpi_card("Total de Propostas", total_p, mono=False, color=COR_ACENTO), unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(kpi_card("Tamanho do Banco", f"{db_size:.1f} KB", mono=False, color=COR_PRIMARIA), unsafe_allow_html=True)

        if not df_inf.empty:
            df_inf["created_at"] = pd.to_datetime(df_inf["created_at"])
            st.markdown(f"<br><p>Primeiro: <strong style='color:#e5e7eb;'>{df_inf['created_at'].min().strftime('%d/%m/%Y')}</strong> &nbsp; \u00daltimo: <strong style='color:#e5e7eb;'>{df_inf['created_at'].max().strftime('%d/%m/%Y')}</strong></p>", unsafe_allow_html=True)

        st.markdown('<div class="pro-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="pro-alert warning">\u26a0 A a\u00e7\u00e3o abaixo remove TODOS os dados permanentemente.</div>', unsafe_allow_html=True)
        if st.checkbox("Confirmo que desejo apagar todos os dados"):
            if st.button("\U0001f5d1 Limpar Banco de Dados", use_container_width=True):
                with get_db() as conn:
                    conn.execute("DELETE FROM proposta_equipe")
                    conn.execute("DELETE FROM proposta_terceiros")
                    conn.execute("DELETE FROM propostas")
                st.success("Banco limpo."); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="pro-card"><div class="pro-card-header"><span class="pro-card-title">Encargos CLT (refer\u00eancia)</span></div>', unsafe_allow_html=True)
    df_enc = pd.DataFrame([
        {"Encargo": k.replace("_"," ").title(), "Percentual (%)": f"{float(v):.4f}"}
        for k,v in ENCARGOS_FPA.items()
    ])
    df_enc.loc[len(df_enc)] = {"Encargo": "FATOR MULTIPLICADOR TOTAL", "Percentual (%)": f"{float(FATOR_ENCARGOS):.6f}\u00d7"}
    st.dataframe(df_enc, hide_index=True, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ==================== SIDEBAR + MAIN ====================

LOGO_URL_SIDEBAR = "https://cdn.prod.website-files.com/65c2dcb4330facd527e06bdd/6619a7597575e945de440959_brivia_group.svg"

PAGES = {
    "dashboard":     ("\U0001f3e0", "Dashboard",      page_dashboard),
    "nova_proposta": ("\u2795",     "Nova Proposta",   page_nova_proposta),
    "propostas":     ("\U0001f4cb", "Propostas",       page_propostas),
    "analytics":     ("\U0001f4ca", "Analytics",       page_analytics),
    "importar":      ("\U0001f4e4", "Importar",        page_importar),
    "configuracoes": ("\u2699\ufe0f","Configura\u00e7\u00f5es", page_config),
}


def render_sidebar():
    with st.sidebar:
        st.markdown(f"""
<div style="padding:24px 20px 16px 20px;border-bottom:1px solid rgba(255,255,255,0.05);margin-bottom:8px;">
  <img src="{LOGO_URL_SIDEBAR}" style="height:26px;filter:brightness(0) invert(1);display:block;margin-bottom:8px;">
  <span style="font-size:0.65rem;text-transform:uppercase;letter-spacing:1.5px;color:rgba(255,255,255,0.22);">Pricing PRO \u00b7 v4.0</span>
</div>""", unsafe_allow_html=True)

        current = st.session_state.get("pagina","dashboard")
        for key, (icon, label, _) in PAGES.items():
            if key == current:
                st.markdown(
                    f'<div style="padding:0 12px 2px 12px;"><div style="background:rgba(197,143,61,0.12);border-left:2px solid #c58f3d;border-radius:8px;padding:10px 14px;font-size:0.85rem;">'
                    f'<span style="color:#c58f3d;">{icon}\u00a0\u00a0{label}</span></div></div>',
                    unsafe_allow_html=True,
                )
            else:
                if st.button(f"{icon}\u00a0\u00a0{label}", key=f"nav_{key}", use_container_width=True):
                    st.session_state.pagina = key
                    # reset wizard when navigating away
                    if key != "nova_proposta":
                        st.session_state.draft = None
                    st.rerun()

        st.markdown("""
<div style="position:absolute;bottom:20px;left:0;right:0;padding:12px 20px;border-top:1px solid rgba(255,255,255,0.05);">
  <span style="font-size:0.68rem;color:rgba(255,255,255,0.18);">Brivia Group \u00a9 2025</span>
</div>""", unsafe_allow_html=True)


def main():
    init_db()
    seed_data()
    inject_css()

    if "pagina" not in st.session_state:
        st.session_state.pagina = "dashboard"

    render_sidebar()

    page_key     = st.session_state.get("pagina","dashboard")
    _, _, page_fn = PAGES.get(page_key, PAGES["dashboard"])
    page_fn()


if __name__ == "__main__":
    main()
